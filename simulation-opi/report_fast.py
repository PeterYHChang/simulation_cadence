import os
import re
import shutil
import uvicorn
import logging
import datetime
import base64
import pdfplumber
import pandas as pd
import numpy as np
from openpyxl.drawing.image import Image
from typing import Optional,Union
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from fastapi.middleware.cors import CORSMiddleware
from fastapi.openapi.utils import get_openapi
from fastapi import FastAPI, HTTPException,Form
from fastapi.responses import JSONResponse, FileResponse
from openpyxl.utils import get_column_letter
from package import export_report,utils

app = FastAPI()
program = 'export_fast'
logpath= os.getcwd()+'\\report_fast'
if not os.path.isdir(logpath):
    os.mkdir(logpath)
logFormatter = '%(asctime)s [%(levelname)s] <%(funcName)s> : %(message)s'
dateFormat = '%Y-%m-%d %H:%M:%S'

logging.basicConfig(filename=os.path.join(logpath,program +'.log'),
                    format=logFormatter, datefmt=dateFormat, level=logging.DEBUG)

logging.info(program)
logging.info("[Start]")
origins = ["*"]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,  # 允許訪問的域名
    allow_credentials=True,
    allow_methods=["*"],  # 允許的 HTTP 請求方法
    allow_headers=["*"],  # 允許的 HTTP 請求標頭
)


def custom_openapi():
    if app.openapi_schema:
        return app.openapi_schema
    openapi_schema = get_openapi(
        title="Export OPI report",
        version="1.0.0",
        description="Export OPI report",
        routes=app.routes,
    )
    openapi_schema["info"]["x-logo"] = {
        "url": "https://fastapi.tiangolo.com/img/logo-margin/logo-teal.png"
    }
    app.openapi_schema = openapi_schema
    return app.openapi_schema

@app.get("/")
async def read_root():
    return {"Hello": "World"}

@app.post("/upload-report/", tags=["產生 Excel 報告"])
async def upload_report(
    project_name: str = Form(...,description="Project name"),
    project_code: str = Form(...,description="Project code"),
    cpu_info: str = Form(...,description="CPU info ex.Intel MTL-U Type3"),
    pcb_no: str = Form(..., description="PCB number"),
    pcb_version: str = Form(..., description="PCB version"),
    net_keyword: str = Form(...,description="Net Keyword, ex.Intel:VCCCORE/ QC:VREG_L17B/ AMD:VDDCR"),
    ori_htm_report:str  = Form(..., description="Original htm Report Name ex.Original_Simulation_Report_VREG_APC0.htm"),
    opt_htm_report: Optional[str]  = Form(None, description="[Optional]Optimize htm Report Name ex.OptimizePI_Report_VREG_APC0.htm"),
    report_path: str = Form(...,description="Orignal/Optimize htm Report Path ex. G:\\simulation-opi_20241216\\output\\20250411_1744339978\\VDDCR"),
    dkdf_path: str = Form(..., description="DKDF File Path ex.G:\\simulation-opi_20241216\\temp\\20250411_1744339978\\243078-SA_DKDF.xlsx"),
    qcm_mapping_table_path: Optional[str]  = Form(None, description="QCM mapping table path to read R/L value.")
):
    """
    ## 操作說明:
    提交表單中的各項資料以進行報告的處理。
    ## 參數說明:
        - project_name: 項目的名稱。
        - project_code: 項目的代碼。
        - cpu_info: CPU 資訊描述，例如 Intel MTL-U Type3。
        - pcb_no: PCB 的編號，例如 223117。
        - pcb_version: PCB 的版本，例如 SA。
        - net_keyword: NetName 關鍵字，例如 Intel:VCCCORE, QC:VREG_L17B, AMD:VDDCR。
        - ori_htm_report: Original HTML 報告的名稱。
        - opt_htm_report: [選填] Optimized HTML 報告的名稱。
        - report_path: Original/Optimized HTML 報告路徑。***(一定要在Sim5)***
        - dkdf_path: DKDF 檔案路徑。***(一定要在Sim5)***
        - qcm_mapping_table_path:[選填] QCM mapping table路徑，用於讀取 R/L 值。***(一定要在Sim5)***
    ## Error:
        HTTPException: 400, No model file
        HTTPException: 400,Cannot Found File
        HTTPException: 400,Any Error

    ## 成功結果:
        請複製以下的 excel_download_url，填入"下載 Excel 報告" 的 excel_report_path 進行下載
        dict:{
            "message": "Files uploaded successfully",
            "pcb_no": pcb_no,
            "pcb_version": pcb_version,
            "excel_download_url": "report_path to download"
            }
    """
    logging.info(f"Get Information-> {project_name=} {project_code=} {cpu_info=} {pcb_no=} {pcb_version=} {net_keyword=}\n {ori_htm_report=} {opt_htm_report=} \n {report_path=} \n{dkdf_path=}\n{qcm_mapping_table_path}")
    try:
        # ------------------------------------ Check Information -------------------------------#
        if qcm_mapping_table_path =='string':
            qcm_mapping_table_path = None
        if opt_htm_report =='string':
            opt_htm_report = None
        # ------------ check model ---------------- #
        vendor,cpu_name,platform,cpu_type,_ = utils.read_CPU_info(cpu_info)
        # 讀取 Platform 的 CPU_conig
        cpu_config_dir=os.path.join(os.getcwd(), 'CPU_config',f"{vendor}_{cpu_name}.json")
        cpu_config = utils.__read_config_file(cpu_config_dir)
        join_name ="_".join([c for c in [cpu_name,platform,cpu_type] if c])
        model_path = os.path.join(os.getcwd(),'Lib',vendor,cpu_name,platform,cpu_type,cpu_config[f"{join_name}_model_path"])
        vrm_path = os.path.join(os.getcwd(),'Lib',vendor,cpu_name,platform,cpu_type,cpu_config["vrm"])
        if not os.path.exists(model_path):
            logging.error(f"No model file :{model_path}")
            raise HTTPException(status_code=400, detail=f"No model file :{model_path}")
        res = check_report_rawdata(report_path,ori_htm_report,opt_htm_report,dkdf_path)
        if res:
            raise HTTPException(status_code=400,detail=f"Cannot Found File {res}")
        # ====================================================================================== #
        ori_htm_path= os.path.join(report_path,ori_htm_report)
        opt_htm_path= os.path.join(report_path,opt_htm_report) if opt_htm_report else None
        # 讀取檔案內容 (這裡只是示範，實際應用可能需要保存檔案或處理內容)
        ori_content = read_htm_report(ori_htm_path)
        opt_content = read_htm_report(opt_htm_path)
        logging.info("Successful to Read HTM Report.")
        # -------- Create Output File and Save html and rowdata ------------ #
        now = datetime.datetime.now()
        output_path = os.path.join(os.getcwd(),'report_fast',project_name+f"_{now.year}{now.month}{now.day}{now.hour}{now.minute}{now.second}")
        os.mkdir(output_path)
        logging.info(f"Create {output_path=}")
        df_pdn =pd.DataFrame({})
        if qcm_mapping_table_path:
            df_pdn = get_qcm_mapping_tbl_data(qcm_mapping_table_path)
            logging.info("Read QC Mapping Table")

        info_dict={
                'customer':"",
                'project_name':project_name,
                'project_code':project_code,
                'platform':platform,
                'product':"",
                'pcbno':pcb_no,
                'pcb_version':pcb_version,
                'stackup_no':"",
                'date':"",
                'power_rail':"",
                'model':"",
                'ref_doc':os.path.basename(model_path),
                'brd':"",
                'other_file':"",
                'schematic':"",
                'net_name_1':"",
                'net_name_2':"",
            }

        report_path,_ = fullfil_report(net_keyword,info_dict,cpu_info,dkdf_path,model_path,ori_content,opt_content,ori_htm_path,output_path,vrm_path,df_pdn)

        # 處理上傳的檔案和表單資料
        # 這裡只是回傳基本資訊，實際應用中可能需要保存檔案或進行其他處理
        response_data = {
            "message": "Files uploaded successfully",
            "pcb_no": pcb_no,
            "pcb_version": pcb_version,
            "excel_download_url": f"{report_path}"
        }
        
        return JSONResponse(content=response_data)
    except Exception as e:
        logging.error(f"error: {e}")
        return HTTPException(status_code=400,detail=f"error: {e}")

# 添加一個專門用於下載 Excel 的端點
@app.get("/download-excel/", tags=["下載 Excel 報告"])
async def download_excel(
    excel_report_path: str
    ):
    """
    ## 操作說明: 提供已存在 Excel 報告的路徑來下載報告。
    ## 參數說明:
        -excel_report_path: 需要下載的 Excel 報告的路徑。
        ex. G:\\\\simulation-opi_20241216\\\\report_fast\\\\string_2025424104140\\\\string_223117-SA_PDN Simulation Result_2025-04-24.xlsx
    ## Error:
        HTTPException: 404, Excel file not found
    ## 成功結果:
        200,"Download file" (藍字，可點擊)
    """
    # 檢查檔案是否存在
    if not os.path.exists(excel_report_path):
        logging.error(f"Excel file not found in {excel_report_path}")
        raise HTTPException(status_code=404, detail=f"Excel file not found : {excel_report_path}")
    logging.error(f"Get excel file in {excel_report_path}")
    # 返回檔案
    return FileResponse(
        path=excel_report_path,
        filename=f"{os.path.basename(excel_report_path)}",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# --------------------------- Function Area ---------------------------------- #

def get_qcm_mapping_tbl_data(qcm_mapping_table_path:str)->pd.DataFrame:
    """Getting Qualcomm Mapping table as a Dataframe 

    Args:
        qcm_mapping_table_path (str): Mapping table path

    Raises:
        NameError: Read QC Mapping Table Error: {error message}

    Returns:
        pd.DataFrame: Information of mapping table
    """
    try:
        df_pdn = pd.DataFrame()  # 初始化為一個空的 DataFrame
        pdn_data = pd.read_excel(qcm_mapping_table_path, sheet_name=None)
        for pdn_sheet,df in pdn_data.items():
            columns_to_read=['CPU Type','Regulator/Power Rail','Layout Power Rail','Power Domain','Positive Pins','Negative Pins','Max Impedance R','Max Impedance L']
            if 'Component' in df.columns:
                columns_to_read.append('Component')
                df['sheet_name']= pdn_sheet
                columns_to_read.append('sheet_name')
                df_pdn = pd.concat([df_pdn, df[columns_to_read]], ignore_index=True)
            else:
                df['sheet_name'] = pdn_sheet
                df['Component'] = np.nan
                columns_to_read.append('sheet_name')
                columns_to_read.append('Component')
                df_pdn = pd.concat([df_pdn, df[columns_to_read]], ignore_index=True)
        return df_pdn
    except Exception as e:
        raise NameError(f"Read QC Mapping Table Error:/n {e}")

def read_htm_report(path:str)->str:
    if not path:
        return ""
    with open(path, 'r', encoding='utf-8') as file:
        content = file.read()
    return content

def check_report_rawdata(report_path:str,ori_htm_report:str,opt_htm_report:str,dkdf_path:str)-> Union[None,str] :
    # 檢查 Original/ Optimize report/ PI_Network_Parameters資料夾有沒有在路徑中。
    # PS Optimize report 可能為 None
    list_file = os.listdir(report_path)
    check_list = [ori_htm_report,"OPI_Network_Parameters"]
    if not os.path.exists(dkdf_path):
        logging.info(f"Cannot Found File {dkdf_path=}")
        return dkdf_path
    if opt_htm_report:
        check_list.append(opt_htm_report)
    for name in check_list:
        if name not in list_file:
            logging.info(f"Cannot Found File {name=}")
            return os.path.join(report_path,name)
    return None

def get_cheme_num(opt_soup:BeautifulSoup)->str:
    """Getting the optimized raw date scheme number from optimized htm report after the title,Cost/Area/Number of Decaps/Number of Type\s+of\s+Capacitor Information.

    Args:
        opt_soup (BeautifulSoup): Optimized soup

    Raises:
        KeyError: Missing OPT Scheme Number Info
        KeyError: Missing OPT Scheme Number Info, error :/n {error message}

    Returns:
        str: scheme number
    """    
    try:
        TARGET_PATTERN = r"Cost/Area/Number\s+of\s+Decaps/Number\s+of\s+Type\s+of\s+Capacitor\s+Information"
        # 找到target_pattern 位置
        target_element = opt_soup.find(string=re.compile(TARGET_PATTERN))
        tables  = target_element.find_all_next('table')
        if tables:
            # 提出target_pattern 後所有表格
            html_content = f"""
                            {tables}
                            """
            list_table=pd.read_html(html_content)
            df_scheme=list_table[0]
            opt_scheme = df_scheme[df_scheme[df_scheme.keys()[0]].str.contains('Optimized Scheme',na=False)].reset_index(drop=True).loc[0][0]
            num_scheme = str(opt_scheme.split(' ')[-1])
            logging.info(f"Get OPT Scheme Number : {num_scheme}")
            return num_scheme
        raise KeyError("Missing OPT Scheme Number Info")
    except Exception as e:
        logging.error(f"Missing OPT Scheme Number Info, error :/n {e}")
        raise KeyError(f"Missing OPT Scheme Number Info, error :/n {e}")

def fullfil_report(net:str,info_dict:dict,cpu_info:str,dkdf_path:str,model_path:str,ori_content:str,opt_content:str,ori_htm_path:str,output_path:str,vrm_path:str,df_pdn:pd.DataFrame)-> tuple[str, str]:
    report_path = os.path.dirname(ori_htm_path)
    pcbno = info_dict['pcbno']
    stage = info_dict['pcb_version']
    project_name = info_dict['project_name']
    platform_info=cpu_info.split(' ')
    vendor=platform_info[0]
    report_config=export_report.__read_report_config()
    hide_sheet=report_config[vendor].get('hide_sheet',"")
    b_hide_sheet=False
    if hide_sheet.lower() =='yes': b_hide_sheet=True
    # Qualcomm OPT HTM report 無 Non-optimized Capacitor Summary 要用template_file_NOPT 的report
    template_file_pass=report_config[vendor].get('report_template_pass',"")
    template_file_fail=report_config[vendor].get('report_template_fail',"")
    # Qualcomm OPT HTM report 有 Non-optimized Capacitor Summary 要用template_file_NOPT 的report
    template_file_non_opt_pass=report_config[vendor].get('NOPT_report_template_pass',"")
    template_file_non_opt_fail=report_config[vendor].get('NOPT_report_template_fail',"")
    template_file_pass=os.path.join(os.getcwd(), 'report_template',template_file_pass)
    template_file_fail=os.path.join(os.getcwd(), 'report_template',template_file_fail)
    template_file_non_opt_pass=os.path.join(os.getcwd(), 'report_template',template_file_non_opt_pass)
    template_file_non_opt_fail=os.path.join(os.getcwd(), 'report_template',template_file_non_opt_fail)
    # ! cost report
    template_file_cost=report_config[vendor].get('cost_template',"")
    if not template_file_cost:
        logging.info("NO_COST_REPORT_TEMPLATE")
        raise NameError('NO_COST_REPORT_TEMPLATE')
    template_file_cost=os.path.join(os.getcwd(), 'report_template',template_file_cost)
    if not os.path.exists(template_file_cost):
        logging.info("task _id: NO_COST_REPORT_TEMPLATE")
        raise FileNotFoundError('NO_COST_REPORT_TEMPLATE')

    opt_soup = BeautifulSoup(opt_content, 'html.parser')
    logging.info("Get OPT_content and Parse.")
    ori_soup = BeautifulSoup(ori_content, 'html.parser')
    logging.info("Get ORI_content and Parse.")
    # --------------------- To find OPT result is Pass/ Fail ----------------------- #
    result=''
    df_opt_status,error_id=export_report.get_target_table(r'Comparing\s+with\s+Target',opt_soup)
    if not df_opt_status.empty:
        column_values = df_opt_status.iloc[1:, 2]
        if  all(value == 'Fail' for value in column_values):
            result = "Fail"
        else:
            result = "Pass"
    logging.info(f"Get OPT Status {result=},{error_id=}")
    # --------------- Create report template ------------- #
    if result =='Pass':
        report_template=template_file_pass
    elif result =='Fail':
            report_template=template_file_fail
    else:
        report_template=template_file_fail

    current_time = datetime.datetime.now()
    formatted_time = current_time.strftime('%Y-%m-%d')
    template_file = load_workbook(report_template)
    template_file_cost=load_workbook(template_file_cost)
    report_filename=f"{project_name}_{pcbno}-{stage}_PDN Simulation Result_{formatted_time}.xlsx"
    cost_report_filename=f"{project_name}_{pcbno}-{stage}_cost_{formatted_time}.xlsx"
    template_file.save(report_filename)
    template_file_cost.save(cost_report_filename)
    # --------------------------------------------------------- ORI Start ------------------------------------------------------ #
    logging.info('ORI report reading..... ')
    #### ---- 讀出 Net pair table
    df_netname,error_id = export_report.use_span_id_to_find_table('IndexByNet',ori_soup)
    list_net_pair=df_netname.iloc[1:,0].tolist()
    list_net=[]
    for netpair in list_net_pair:
        list_net+=[netpair.split(':')[0]]
        logging.info(f"Get *ORI_Net pair* {netpair}, {error_id=}")
    #### ---- Report_stackup
    # 尋找 1.2 Board Stackup 的 table
    # target_pattern = r"1\.2\s+Board\s+Stackup"
    # 以<span id="BoardStackup">查找
    df_stackup,error_id = export_report.use_span_id_to_find_table('BoardStackup',ori_soup)
    logging.info(f"Get *Board Stackup* Table,{error_id=}")

    #### ---- ORI_SummaryCap
    df_ori_summary,error_id = export_report.keyword_target_get_summary_table (r'Capacitor\s+Summary','QTY',ori_soup,list_net_pair)
    logging.info(f"Get *ORI_SummaryCap* Table,{error_id=}")

    #### ---- ORI_TotalCap
    df_ori_table,error_id = export_report.keyword_target_get_summary_table (r'Capacitor\s+Placement','REFDES',ori_soup,list_net_pair)
    logging.info(f"Get *ORI_TotalCap* Table,{error_id=}")
    #### ---- ORI_Status
    df_ori_status,error_id=export_report.get_target_table(r"Comparing\s+with\s+Target",ori_soup)
    logging.info(f"Get *ORI_Status* Table,{error_id=}")

    # FAST API 不需要 ori_result=''
    # if not df_ori_status.empty:
    #     column_values = df_ori_status.iloc[1:, 2]
    #     ori_result='Fail'
    #     if all(value == 'Fail' for value in column_values):
    #         ori_result='Fail'
    #     else:
    #         ori_result='Pass'

    #### ---- ORI_RowData
    if vendor.upper() == 'INTEL':
        ori_raw_data,error_id = export_report.get_rawdata(vendor,"",pd.DataFrame(),report_path,ori_soup)
    elif vendor.upper() == 'AMD':
        ori_raw_data,error_id = export_report.get_rawdata(vendor,list_net[0],pd.DataFrame(),report_path,ori_soup)
    elif vendor.upper() == 'QUALCOMM':
        ori_raw_data,error_id = export_report.get_rawdata(vendor,list_net[0],df_ori_status,report_path,ori_soup)

    logging.info(f"Get *ORI_RowData* Table,{error_id=}")
    #### ---- ORI_CapLocations
    src_list,error_id = export_report.get_image(r'Device\s+and\s+Capacitor\s+Locations',ori_soup)
    logging.info(f"Get *ORI_CapLocations*,{error_id=}")
    # 輸出所有的ORI src內容
    img_list=[]
    for num in range(len(list_net_pair)*2):
        image_data = base64.b64decode(src_list[num].split(',')[-1])
        print(image_data)
        # 將圖片保存為jpg格式
        file_name=os.path.join(output_path,f'ori_location_{num}.jpg')
        with open(file_name,"wb") as f:
            f.write(image_data)
        img_list.append(file_name)
    #### ---- Result_PDN 結果圖
    src_list,error_id = export_report.get_image(r'Impedances\s+at\s+Devices',ori_soup)
    logging.info(f"Get *Result_PDN 結果圖*,{error_id=}")
    # 輸出所有的ORI src內容
    ori_img_list=[]
    for num in range(len(src_list)):
        if vendor.upper()=='QUALCOMM':
            image_data = base64.b64decode(src_list[num].split(',')[-1])
            # 將圖片保存為jpg格式
            file_name=os.path.join(output_path,f'ori_output_{num}.jpg')
            with open(file_name,"wb") as f:
                f.write(image_data)
            ori_img_list.append(file_name)
            continue
        # Intel/ AMD 才執行以下
        if num==0:
            image_data = base64.b64decode(src_list[num].split(',')[-1])
            # 將圖片保存為jpg格式
            file_name=os.path.join(output_path,f'ori_output_{num}.jpg')
            with open(file_name,"wb") as f:
                f.write(image_data)
            ori_img_list.append(file_name)
            break

    # --------------------------------------------------------- OPT Start ------------------------------------------------------ #
    logging.info('OPT report reading..... ')
    #### ---- 讀出 OPT Net pair table
    df_netname,error_id = export_report.use_span_id_to_find_table('IndexByNet',opt_soup)
    if not df_netname.empty:
        logging.info(f"Get *OPT_Net pair*, {error_id=}")
        list_net_pair=df_netname.iloc[1:,0].tolist()
        list_net=[]
        for netpair in list_net_pair:
            list_net+=[netpair.split(':')[0]]
            logging.info(f"Get *OPT_Net pair* {list_net}, {error_id=}")

    df_opt_totalcap = pd.DataFrame()
    df_nopt_totalcap = pd.DataFrame()
    df_opt_summuary=pd.DataFrame()
    opt_raw_data=pd.DataFrame()
    df_rldate=pd.DataFrame()

    if not result:
        logging.info("Missing : *OPT_result*")
    elif result in ['Pass','Fail']:
        #### ---- OPT_SummaryCap
        target_patterns = [
                            r'Non-optimized\s+Capacitors\s+Summary',
                            r'Optimized\s+Capacitor\s+Summary'
                            ]
        
        for _, target_pattern in enumerate(target_patterns):
            df_table,error_id = export_report.keyword_target_get_summary_table(target_pattern,'QTY',opt_soup,list_net_pair)
            logging.info(f'Get *{target_pattern}* Table, {error_id=}')
            #! 有些報告沒有 Non-optimized 的 table，但 Summuary table 不受影響，有就輸出
            if df_opt_summuary.empty:
                df_opt_summuary = df_table
                continue
            df_process = df_table.drop(0, axis=0)
            df_process.reset_index(drop=True, inplace=True)
            df_opt_summuary = pd.concat([df_opt_summuary, df_process], ignore_index=True)
        #### ---- OPT_TotalCap
        # if result == 'Pass' or result == '':
        target_patterns = [
                            r'Non-optimized\s+Capacitors\s+Placement',
                            r'Optimized\s+Capacitor\s+Placement'
                        ]
        for _, target_pattern in enumerate(target_patterns):
            df_table,error_id = export_report.keyword_target_get_summary_table(target_pattern,'REFDES',opt_soup,list_net_pair)
            logging.info(f'Get *{target_pattern}* Table, {error_id=}')
            #! 有 Non-optimized 的 table，則將 df_NOPT_TotalCap 放入資料，否則為空 dataframe 輸出
            if target_pattern == r'Non-optimized\s+Capacitors\s+Placement':
                df_nopt_totalcap = df_table
            #! 若有Non-optimized 的 table，df_OPT_TotalCap 開頭要先放Non-optimized 的結果，在接上 Optimized 的結果
            if df_opt_totalcap.empty:
                df_opt_totalcap = pd.concat([df_opt_totalcap, df_table], ignore_index=True)
                continue
            df_process = df_table.drop(0, axis=0)
            df_process.reset_index(drop=True, inplace=True)
            df_opt_totalcap = pd.concat([df_opt_totalcap, df_process], ignore_index=True)
            logging.info(f"Get *{target_pattern}* Table")

        #### ---- OPI_15%
        # * 讀取圖檔 * #
        src_list,error_id = export_report.get_image(r'Impedances\s+at\s+Devices',opt_soup)
        # 輸出所有的ORI src內容
        opt_img_list=[]
        for num in range(len(src_list)):
            image_data = base64.b64decode(src_list[num].split(',')[-1])
            logging.info(f"Get *OPI_15%*, {error_id=}")
            # 將圖片保存為jpg格式
            file_name=os.path.join(output_path,f'OPT_output_{num}.jpg')
            with open(file_name,"wb") as f:
                f.write(image_data)
            opt_img_list.append(file_name)
            logging.info("OPT Image Output")
        #### ---- OPT_RawData for Qualcomm
        num_schem = get_cheme_num(opt_soup)
        if vendor.upper == 'INTEL':
            opt_raw_data,error_id = export_report.get_rawdata(vendor,"",pd.DataFrame(),report_path,opt_soup,'Scheme',num_schem)
        elif vendor.upper == 'AMD':
            opt_raw_data,error_id = export_report.get_rawdata(vendor,list_net[0],pd.DataFrame(),report_path,opt_soup,'Scheme',num_schem)
        elif vendor.upper == 'QUALCOMM':
            opt_raw_data,error_id = export_report.get_rawdata(vendor,list_net[0],df_ori_status,report_path,opt_soup,'Scheme',num_schem)
            df_rldate,rl_error_id = export_report.get_qc_rl_data(net,df_opt_status,opt_soup,df_pdn=df_pdn)
            logging.info(f"Get *QCM_RL_date*, {rl_error_id=}")

    # * 找出路徑中pdf的檔名
    model_pdf_path,info_dict['model'] = export_report.get_model_pdf(net,model_path)
    # * report date
    report_create_date=export_report.__file_crate_time(ori_htm_path)
    #### ---- 輸出 Report_info            
    ver=export_report.__read_resource_profile_log(os.path.join(report_path,'ResourceProfile.log'))
    export_report.__table(pd.DataFrame([{'Date':report_create_date,'Version':ver}]),report_filename,'Report_info',b_hide_sheet)
    logging.info("Output Report_info")
    #### ---- 輸出 stackup_dkdf
    export_report.__stackup(dkdf_path,report_filename,b_hide_sheet)
    logging.info('Output stackup_dkdf')
    #### ---- 輸出 Info
    df_info = pd.DataFrame([info_dict])
    export_report.__table(df_info,report_filename,'Web_info1',b_hide_sheet)
    logging.info('Output Info')
    #### ---- 輸出 Report_ModelVRM
    if os.path.basename(vrm_path) !='':
        export_report.__read_vrm(vrm_path,net,report_filename,'Report_ModelVRM',b_hide_sheet)
        logging.info('Output Report_ModelVRM')
    #### ---- 輸出 Report_stackup
    export_report.__table(df_stackup,report_filename,'Report_stackup',b_hide_sheet)
    logging.info('Output Report_stackup')
    #### ---- 輸出 ORI_SummaryCap
    export_report.__table(df_ori_summary,report_filename,'ORI_SummaryCap',b_hide_sheet)
    logging.info('Output ORI_SummaryCap')
    #### ---- 輸出 ORI_TotalCap
    export_report.__table(df_ori_table,report_filename,'ORI_TotalCap',b_hide_sheet)
    logging.info('Output ORI_TotalCap')
    #### ---- 輸出 ORI_RowData
    export_report.__table(ori_raw_data,report_filename,'ORI_RowData',b_hide_sheet)
    logging.info('Output ORI_RowData')
    #### ---- 輸出 ORI_Status
    export_report.__table(df_ori_status,report_filename,'ORI_Status',b_hide_sheet)
    logging.info('Output ORI_Status')
    #### ---- 輸出 ORI_CapLocations
    export_report.__image_to_worksheet(img_list,report_filename,'Result_ PDN_Capacitor Locations',False,vendor)
    logging.info('Output ORI_CapLocations')
    #### ---- Result_PDN
    export_report.__result_image_to_worksheet(ori_img_list,report_filename,'Result_PDN',False,vendor)
    logging.info('Output ORI result')

    if vendor.upper()=='INTEL':
        #### ---- 讀取 model.pdf
        imag_path=[]
        with pdfplumber.open(model_pdf_path) as pdf:
            for page_num in range(len(pdf.pages)):
                page=pdf.pages[page_num]
                # 找到有 "Impedance at Package" 的 Page
                if "Impedance at Package" not in page.extract_text():
                    logging.info('No Criteria image')
                    continue
                # 有 "Impedance at Package" 的 Page 輸出成jpg
                image = page.to_image()
                image.save(os.path.join(output_path,f'impedance_page_{page_num+1}.jpg'))
                imag_path.append(os.path.join(output_path,f'impedance_page_{page_num+1}.jpg'))

            if imag_path!=[]:
                index_r=8
                index_c=1
                for imag in imag_path:
                    output_file = load_workbook(report_filename)
                    # * get sheet name
                    remove_idex=output_file.get_sheet_by_name('Criteria')
                    output_sheet = remove_idex
                    image_read = Image(imag)
                    image_read.width = 720
                    image_read.height = 350
                    output_sheet.add_image(image_read, f"{get_column_letter(index_c)}{index_r}")
                    output_file.save(report_filename)
                    index_c+=14
                logging.info('Output Criteria image')
            else:
                logging.info('Missing Criteria image')
    if result:
        if vendor.upper()=='INTEL':
            #### ---- 輸出 OPT_Status
            export_report.__table(df_opt_status,report_filename,'OPT_Status',b_hide_sheet)
            logging.info('Output OPT_Status')
            #### ---- Report_Target
            files = os.listdir(model_path)
            key_file = [file for file in files if net in file][0]
            path = os.path.join(model_path,key_file)
            # * 找出路徑中的target(xlsx) (intel 標準)
            files = os.listdir(path)
            target_files = [file for file in files if file.endswith('.xlsx')]
            target_file=target_files[0] # * target file
            export_report.__target(os.path.join(path,target_file),report_filename,'Report_Target',b_hide_sheet)
        if vendor.upper()=='QUALCOMM':
            #### ---- 輸出 Report_Target/QC 輸出 RL 值
            export_report.__table(df_rldate,report_filename,'Report_Target',b_hide_sheet)
            logging.info('Output Report_Target/QC, Output RL value')
        if vendor.upper()=='AMD':
            #### ---- 輸出 OPT_Status
            export_report.__table(df_opt_status,report_filename,'OPT_Status',b_hide_sheet)
            logging.info('Output OPT_Status')
        export_report.__table(df_opt_summuary,report_filename,'OPT_SummaryCap',b_hide_sheet)
        logging.info('Output OPT_SummaryCap')
        #### ---- 輸出 OPT_TotalCap
        export_report.__table(df_opt_totalcap,report_filename,'OPT_TotalCap',b_hide_sheet)
        logging.info('Output OPT_TotalCap')
        #### ---- 輸出 OPT_RowData
        export_report.__table(opt_raw_data,report_filename,'OPT_RowData',b_hide_sheet)
        logging.info('Output OPT_RowData')
    if result == 'Pass':
        #### ---- Result_PDN 
        export_report.__result_image_to_worksheet(opt_img_list,report_filename,'OPI_15%',False,vendor)
        logging.info('Output OPT result')
    logging.info(f'OPT Status:{result}')
    # ! ---- cost report -----------------------------------------------------------
    logging.info('Output cost report')
    #### ---- NetName
    export_report.__table(df_netname,cost_report_filename,'NetName',b_hide_sheet)
    #### ---- ORI_TotalCap
    export_report.__table(df_ori_table,cost_report_filename,'ORI_TotalCap',b_hide_sheet)
    if result:
        #### ---- OPT_Status
        export_report.__table(df_opt_status,cost_report_filename,'OPT_Status',b_hide_sheet)
        #### ---- OPT_TotalCap
        export_report.__table(df_opt_totalcap,cost_report_filename,'OPT_TotalCap',b_hide_sheet)
        #### ---- NOPT_TotalCap
        export_report.__table(df_nopt_totalcap,cost_report_filename,'NOPT_TotalCap',b_hide_sheet)
        logging.info('Output cost report done')

    # ! 讀取cost資料準備進DB
    # REMARK
    # export_report.__read_cost_summary(cost_report_filename)

    # copy Excel 到資料夾位置
    shutil.copy2(report_filename, output_path)
    os.remove(report_filename) # 刪除檔案
    logging.info(f"Copy to :{output_path}")
    shutil.copy2(cost_report_filename, output_path)
    os.remove(cost_report_filename) # 刪除檔案
    return os.path.join(output_path,report_filename),os.path.join(output_path,cost_report_filename)

if __name__ == '__main__':
    # ssl_cert = os.path.join(os.getcwd(), 'STAR_wistron.com.crt')
    # ssl_key = os.path.join(os.getcwd(), 'STAR_wistron.com.key')
    app.openapi = custom_openapi
    # Uncomment for SSL
    # uvicorn.run('report_fast:app',
    #         ssl_keyfile=ssl_key,
    #         ssl_certfile=ssl_cert,
    #         host="0.0.0.0",
    #         port=2516,
    #         reload=True)
    uvicorn.run('report_fast:app',host="0.0.0.0",port=2516, reload=True)  # Ensure the correct module path is used

    # import bson
    # _id = '6808a94d82adf8505cdf0ee5'
    # ### 65cf0785ef6d604367056fdc
    # collect_tc = utils.ConnectToMongoDB(strDBName="Simulation",strTableName= 'OPITaskCtrl_Debug')
    # task_info=collect_tc.find_one({'_id':bson.ObjectId(_id)})
    # # ++++++++++++++++++++++++++++ 測試用 parameter setting +++++++++++++++++++++++++ #
    # idx=0
    # tcl=task_info['tcl'][idx]
    # running_task=task_info
    # export_report.__export_report(_id,tcl,running_task,idx,'auto')

# --------------------------------------- test debug -----------------------------------
    # net_keyword ='VCCCORE'
    # net = net_keyword
    # info_dict = {
    #                 'customer':"",
    #                 'project_name':"",
    #                 'project_code':"project_code",
    #                 'platform':"platform",
    #                 'product':"",
    #                 'pcbno':"pcb_no",
    #                 'pcb_version':"pcb_version",
    #                 'stackup_no':"",
    #                 'date':"",
    #                 'power_rail':"",
    #                 'model':"",
    #                 'ref_doc':"",
    #                 'brd':"",
    #                 'other_file':"",
    #                 'schematic':"",
    #                 'net_name_1':"",
    #                 'net_name_2':"",
    #             }
    # cpu_info = 'Qualcomm SC8480XP-5_3_3_5'
    # dkdf_path = r'G:\simulation-opi\temp\243064-SB_1740564961\stackup_dkdf.xlsx'
    # model_path =""
    # report_path = r'G:\simulation-opi\output\243064-SB_1740564961\3.2DDR PDN\VREG_S1L1_S2L1'
    # ori_htm_report = 'Original_Simulation_Report_VREG_S1L1_S2L1.htm'
    # opt_htm_report = ''
    # ori_htm_path= os.path.join(report_path,ori_htm_report)
    # opt_htm_path= os.path.join(report_path,opt_htm_report) if opt_htm_report else None
    # # 讀取檔案內容 (這裡只是示範，實際應用可能需要保存檔案或處理內容)
    # ori_content = read_htm_report(ori_htm_path)
    # opt_content = read_htm_report(opt_htm_path)
    # output_path = report_path
    # vrm_path=""
    # df_pdn=pd.DataFrame({})
    # report_path = os.path.dirname(ori_htm_path)
    # pcbno = info_dict['pcbno']
    # stage = info_dict['pcb_version']
    # project_name = info_dict['project_name']
    # platform_info=cpu_info.split(' ')
    # vendor=platform_info[0]
    # report_config=export_report.__read_report_config()
    # hide_sheet=report_config[vendor].get('hide_sheet',"")
    # b_hide_sheet=False
    # if hide_sheet.lower() =='yes': b_hide_sheet=True
    # # Qualcomm OPT HTM report 無 Non-optimized Capacitor Summary 要用template_file_NOPT 的report
    # template_file_pass=report_config[vendor].get('report_template_pass',"")
    # template_file_fail=report_config[vendor].get('report_template_fail',"")
    # # Qualcomm OPT HTM report 有 Non-optimized Capacitor Summary 要用template_file_NOPT 的report
    # template_file_non_opt_pass=report_config[vendor].get('NOPT_report_template_pass',"")
    # template_file_non_opt_fail=report_config[vendor].get('NOPT_report_template_fail',"")
    # template_file_pass=os.path.join(os.getcwd(), 'report_template',template_file_pass)
    # template_file_fail=os.path.join(os.getcwd(), 'report_template',template_file_fail)
    # template_file_non_opt_pass=os.path.join(os.getcwd(), 'report_template',template_file_non_opt_pass)
    # template_file_non_opt_fail=os.path.join(os.getcwd(), 'report_template',template_file_non_opt_fail)
    # # ! cost report
    # template_file_cost=report_config[vendor].get('cost_template',"")
    # if not template_file_cost:
    #     logging.info("NO_COST_REPORT_TEMPLATE")
    #     raise NameError('NO_COST_REPORT_TEMPLATE')
    # template_file_cost=os.path.join(os.getcwd(), 'report_template',template_file_cost)
    # if not os.path.exists(template_file_cost):
    #     logging.info("task _id: NO_COST_REPORT_TEMPLATE")
    #     raise FileNotFoundError('NO_COST_REPORT_TEMPLATE')

    # opt_soup = BeautifulSoup(opt_content, 'html.parser')
    # logging.info("Get OPT_content and Parse.")
    # ori_soup = BeautifulSoup(ori_content, 'html.parser')
    # logging.info("Get ORI_content and Parse.")
    # # --------------------- To find OPT result is Pass/ Fail ----------------------- #
    # result=''
    # df_opt_status,error_id=export_report.get_target_table(r'Comparing\s+with\s+Target',opt_soup)
    # if not df_opt_status.empty:
    #     column_values = df_opt_status.iloc[1:, 2]
    #     if  all(value == 'Fail' for value in column_values):
    #         result = "Fail"
    #     else:
    #         result = "Pass"
    # logging.info(f"Get OPT Status {result=},{error_id=}")
    # # --------------- Create report template ------------- #
    # if result =='Pass':
    #     report_template=template_file_pass
    # elif result =='Fail':
    #         report_template=template_file_fail
    # else:
    #     report_template=template_file_fail

    # current_time = datetime.datetime.now()
    # formatted_time = current_time.strftime('%Y-%m-%d')
    # template_file = load_workbook(report_template)
    # template_file_cost=load_workbook(template_file_cost)
    # report_filename=f"{project_name}_{pcbno}-{stage}_PDN Simulation Result_{formatted_time}.xlsx"
    # cost_report_filename=f"{project_name}_{pcbno}-{stage}_cost_{formatted_time}.xlsx"
    # template_file.save(report_filename)
    # template_file_cost.save(cost_report_filename)
    # # --------------------------------------------------------- ORI Start ------------------------------------------------------ #
    # logging.info('ORI report reading..... ')
    # #### ---- 讀出 Net pair table
    # df_netname,error_id = export_report.use_span_id_to_find_table('IndexByNet',ori_soup)
    # list_net_pair=df_netname.iloc[1:,0].tolist()
    # list_net=[]
    # for netpair in list_net_pair:
    #     list_net+=[netpair.split(':')[0]]
    #     logging.info(f"Get *ORI_Net pair* {netpair}, {error_id=}")
    # #### ---- Report_stackup
    # # 尋找 1.2 Board Stackup 的 table
    # # target_pattern = r"1\.2\s+Board\s+Stackup"
    # # 以<span id="BoardStackup">查找
    # df_stackup,error_id = export_report.use_span_id_to_find_table('BoardStackup',ori_soup)
    # logging.info(f"Get *Board Stackup* Table,{error_id=}")

    # #### ---- ORI_SummaryCap
    # df_ori_summary,error_id = export_report.keyword_target_get_summary_table (r'Capacitor\s+Summary','QTY',ori_soup,list_net_pair)
    # logging.info(f"Get *ORI_SummaryCap* Table,{error_id=}")

    # #### ---- ORI_TotalCap
    # df_ori_table,error_id = export_report.keyword_target_get_summary_table (r'Capacitor\s+Placement','REFDES',ori_soup,list_net_pair)
    # logging.info(f"Get *ORI_TotalCap* Table,{error_id=}")
    # #### ---- ORI_Status
    # df_ori_status,error_id=export_report.get_target_table(r"Comparing\s+with\s+Target",ori_soup)
    # logging.info(f"Get *ORI_Status* Table,{error_id=}")

    # # FAST API 不需要 ori_result=''
    # # column_values = df_ori_status.iloc[1:, 2]
    # # ori_result='Fail'
    # # if all(value == 'Fail' for value in column_values):
    # #     ori_result='Fail'
    # # else:
    # #     ori_result='Pass'

    # #### ---- ORI_RowData
    # if vendor.upper() == 'INTEL':
    #     ori_raw_data,error_id = export_report.get_rawdata(vendor,"",pd.DataFrame(),report_path,ori_soup)
    # elif vendor.upper() == 'AMD':
    #     ori_raw_data,error_id = export_report.get_rawdata(vendor,list_net[0],pd.DataFrame(),report_path,ori_soup)
    # elif vendor.upper() == 'QUALCOMM':
    #     ori_raw_data,error_id = export_report.get_rawdata(vendor,list_net[0],df_ori_status,report_path,ori_soup)

    # logging.info(f"Get *ORI_RowData* Table,{error_id=}")
    # #### ---- ORI_CapLocations
    # src_list,error_id = export_report.get_image(r'Device\s+and\s+Capacitor\s+Locations',ori_soup)
    # logging.info(f"Get *ORI_CapLocations*,{error_id=}")
    # # 輸出所有的ORI src內容
    # img_list=[]
    # for num in range(len(list_net_pair)*2):
    #     image_data = base64.b64decode(src_list[num].split(',')[-1])
    #     print(image_data)
    #     # 將圖片保存為jpg格式
    #     file_name=os.path.join(output_path,f'ori_location_{num}.jpg')
    #     with open(file_name,"wb") as f:
    #         f.write(image_data)
    #     img_list.append(file_name)
    # #### ---- Result_PDN 結果圖
    # src_list,error_id = export_report.get_image(r'Impedances\s+at\s+Devices',ori_soup)
    # logging.info(f"Get *Result_PDN 結果圖*,{error_id=}")
    # # 輸出所有的ORI src內容
    # ori_img_list=[]
    # for num in range(len(src_list)):
    #     if vendor.upper()=='QUALCOMM':
    #         image_data = base64.b64decode(src_list[num].split(',')[-1])
    #         # 將圖片保存為jpg格式
    #         file_name=os.path.join(output_path,f'ori_output_{num}.jpg')
    #         with open(file_name,"wb") as f:
    #             f.write(image_data)
    #         ori_img_list.append(file_name)
    #         continue
    #     # Intel/ AMD 才執行以下
    #     if num==0:
    #         image_data = base64.b64decode(src_list[num].split(',')[-1])
    #         # 將圖片保存為jpg格式
    #         file_name=os.path.join(output_path,f'ori_output_{num}.jpg')
    #         with open(file_name,"wb") as f:
    #             f.write(image_data)
    #         ori_img_list.append(file_name)
    #         break

    # # --------------------------------------------------------- OPT Start ------------------------------------------------------ #
    # logging.info('OPT report reading..... ')
    # #### ---- 讀出 OPT Net pair table
    # df_netname,error_id = export_report.use_span_id_to_find_table('IndexByNet',opt_soup)
    # if not df_netname.empty:
    #     logging.info(f"Get *OPT_Net pair*, {error_id=}")
    #     list_net_pair=df_netname.iloc[1:,0].tolist()
    #     list_net=[]
    #     for netpair in list_net_pair:
    #         list_net+=[netpair.split(':')[0]]
    #         logging.info(f"Get *OPT_Net pair* {list_net}, {error_id=}")

    # df_opt_totalcap = pd.DataFrame()
    # df_nopt_totalcap = pd.DataFrame()
    # df_opt_summuary=pd.DataFrame()
    # opt_raw_data=pd.DataFrame()
    # df_rldate=pd.DataFrame()

    # if not result:
    #     logging.info("Missing : *OPT_result*")
    # elif result in ['Pass','Fail']:
    #     #### ---- OPT_SummaryCap
    #     target_patterns = [
    #                         r'Non-optimized\s+Capacitors\s+Summary',
    #                         r'Optimized\s+Capacitor\s+Summary'
    #                         ]
        
    #     for _, target_pattern in enumerate(target_patterns):
    #         df_table,error_id = export_report.keyword_target_get_summary_table(target_pattern,'QTY',opt_soup,list_net_pair)
    #         logging.info(f'Get *{target_pattern}* Table, {error_id=}')
    #         #! 有些報告沒有 Non-optimized 的 table，但 Summuary table 不受影響，有就輸出
    #         if df_opt_summuary.empty:
    #             df_opt_summuary = df_table
    #             continue
    #         df_process = df_table.drop(0, axis=0)
    #         df_process.reset_index(drop=True, inplace=True)
    #         df_opt_summuary = pd.concat([df_opt_summuary, df_process], ignore_index=True)
    #     #### ---- OPT_TotalCap
    #     # if result == 'Pass' or result == '':
    #     target_patterns = [
    #                         r'Non-optimized\s+Capacitors\s+Placement',
    #                         r'Optimized\s+Capacitor\s+Placement'
    #                     ]
    #     for _, target_pattern in enumerate(target_patterns):
    #         df_table,error_id = export_report.keyword_target_get_summary_table(target_pattern,'REFDES',opt_soup,list_net_pair)
    #         logging.info(f'Get *{target_pattern}* Table, {error_id=}')
    #         #! 有 Non-optimized 的 table，則將 df_NOPT_TotalCap 放入資料，否則為空 dataframe 輸出
    #         if target_pattern == r'Non-optimized\s+Capacitors\s+Placement':
    #             df_nopt_totalcap = df_table
    #         #! 若有Non-optimized 的 table，df_OPT_TotalCap 開頭要先放Non-optimized 的結果，在接上 Optimized 的結果
    #         if df_opt_totalcap.empty:
    #             df_opt_totalcap = pd.concat([df_opt_totalcap, df_table], ignore_index=True)
    #             continue
    #         df_process = df_table.drop(0, axis=0)
    #         df_process.reset_index(drop=True, inplace=True)
    #         df_opt_totalcap = pd.concat([df_opt_totalcap, df_process], ignore_index=True)
    #         logging.info(f"Get *{target_pattern}* Table")

    #     #### ---- OPI_15%
    #     # * 讀取圖檔 * #
    #     src_list,error_id = export_report.get_image(r'Impedances\s+at\s+Devices',opt_soup)
    #     # 輸出所有的ORI src內容
    #     opt_img_list=[]
    #     for num in range(len(src_list)):
    #         image_data = base64.b64decode(src_list[num].split(',')[-1])
    #         logging.info(f"Get *OPI_15%*, {error_id=}")
    #         # 將圖片保存為jpg格式
    #         file_name=os.path.join(output_path,f'OPT_output_{num}.jpg')
    #         with open(file_name,"wb") as f:
    #             f.write(image_data)
    #         opt_img_list.append(file_name)
    #         logging.info("OPT Image Output")
    #     #### ---- OPT_RawData for Qualcomm
    #     num_schem = get_cheme_num(opt_soup)
    #     if vendor.upper == 'INTEL':
    #         opt_raw_data,error_id = export_report.get_rawdata(vendor,"",pd.DataFrame(),report_path,opt_soup,'Scheme',num_schem)
    #         df_rldate,rl_error_id = export_report.get_qc_rl_data(net,df_opt_status,opt_soup,df_pdn=df_pdn)
    #         logging.info(f"Get *QC_RL_date*, {rl_error_id=}")
    #     elif vendor.upper == 'AMD':
    #         opt_raw_data,error_id = export_report.get_rawdata(vendor,list_net[0],pd.DataFrame(),report_path,opt_soup,'Scheme',num_schem)
    #     elif vendor.upper == 'QUALCOMM':
    #         opt_raw_data,error_id = export_report.get_rawdata(vendor,list_net[0],df_ori_status,report_path,opt_soup,'Scheme',num_schem)

    # # * 找出路徑中pdf的檔名
    # model_pdf_path,info_dict['model'] = export_report.get_model_pdf(net,model_path)
    # # * report date
    # report_create_date=export_report.__file_crate_time(ori_htm_path)
    # #### ---- 輸出 Report_info            
    # ver=export_report.__read_resource_profile_log(os.path.join(report_path,'ResourceProfile.log'))
    # export_report.__table(pd.DataFrame([{'Date':report_create_date,'Version':ver}]),report_filename,'Report_info',b_hide_sheet)
    # logging.info("輸出 Report_info")
    # #### ---- 輸出 stackup_dkdf
    # export_report.__stackup(dkdf_path,report_filename,b_hide_sheet)
    # logging.info('輸出 stackup_dkdf')
    # #### ---- 輸出 Info
    # df_info = pd.DataFrame([info_dict])
    # export_report.__table(df_info,report_filename,'Web_info1',b_hide_sheet)
    # logging.info('輸出 Info')
    # #### ---- 輸出 Report_ModelVRM
    # if os.path.basename(vrm_path) !='':
    #     export_report.__read_vrm(vrm_path,net,report_filename,'Report_ModelVRM',b_hide_sheet)
    #     logging.info('輸出 Report_ModelVRM')
    # #### ---- 輸出 Report_stackup
    # export_report.__table(df_stackup,report_filename,'Report_stackup',b_hide_sheet)
    # logging.info('輸出 Report_stackup')
    # #### ---- 輸出 ORI_SummaryCap
    # export_report.__table(df_ori_summary,report_filename,'ORI_SummaryCap',b_hide_sheet)
    # logging.info('輸出 ORI_SummaryCap')
    # #### ---- 輸出 ORI_TotalCap
    # export_report.__table(df_ori_table,report_filename,'ORI_TotalCap',b_hide_sheet)
    # logging.info('輸出 ORI_TotalCap')
    # #### ---- 輸出 ORI_RowData
    # export_report.__table(ori_raw_data,report_filename,'ORI_RowData',b_hide_sheet)
    # logging.info('輸出 ORI_RowData')
    # #### ---- 輸出 ORI_Status
    # export_report.__table(df_ori_status,report_filename,'ORI_Status',b_hide_sheet)
    # logging.info('輸出 ORI_Status')
    # #### ---- 輸出 ORI_CapLocations
    # export_report.__image_to_worksheet(img_list,report_filename,'Result_ PDN_Capacitor Locations',False,vendor)
    # logging.info('輸出 ORI_CapLocations')
    # #### ---- Result_PDN
    # export_report.__result_image_to_worksheet(ori_img_list,report_filename,'Result_PDN',False,vendor)
    # logging.info('輸出 ORI result')

    # if vendor.upper()=='INTEL':
    #     #### ---- 讀取 model.pdf
    #     imag_path=[]
    #     with pdfplumber.open(model_pdf_path) as pdf:
    #         for page_num in range(len(pdf.pages)):
    #             page=pdf.pages[page_num]
    #             # 找到有 "Impedance at Package" 的 Page
    #             if "Impedance at Package" not in page.extract_text():
    #                 logging.info('No Criteria image')
    #                 continue
    #             # 有 "Impedance at Package" 的 Page 輸出成jpg
    #             image = page.to_image()
    #             image.save(os.path.join(output_path,f'impedance_page_{page_num+1}.jpg'))
    #             imag_path.append(os.path.join(output_path,f'impedance_page_{page_num+1}.jpg'))

    #         if imag_path!=[]:
    #             index_r=8
    #             index_c=1
    #             for imag in imag_path:
    #                 output_file = load_workbook(report_filename)
    #                 # * get sheet name
    #                 remove_idex=output_file.get_sheet_by_name('Criteria')
    #                 output_sheet = remove_idex
    #                 image_read = Image(imag)
    #                 image_read.width = 720
    #                 image_read.height = 350
    #                 output_sheet.add_image(image_read, f"{get_column_letter(index_c)}{index_r}")
    #                 output_file.save(report_filename)
    #                 index_c+=14
    #             logging.info('輸出 Criteria image')
    #         else:
    #             logging.info('Missing Criteria image')
    # if result:
    #     if vendor.upper()=='INTEL':
    #         #### ---- 輸出 OPT_Status
    #         export_report.__table(df_opt_status,report_filename,'OPT_Status',b_hide_sheet)
    #         logging.info('輸出 OPT_Status')
    #         #### ---- Report_Target
    #         files = os.listdir(model_path)
    #         key_file = [file for file in files if net_keyword in file][0]
    #         path = os.path.join(model_path,key_file)
    #         # * 找出路徑中的target(xlsx) (intel 標準)
    #         files = os.listdir(path)
    #         target_files = [file for file in files if file.endswith('.xlsx')]
    #         target_file=target_files[0] # * target file
    #         export_report.__target(os.path.join(path,target_file),report_filename,'Report_Target',b_hide_sheet)
    #     if vendor.upper()=='QUALCOMM':
    #         #### ---- 輸出 Report_Target/QC 輸出 RL 值
    #         export_report.__table(df_rldate,report_filename,'Report_Target',b_hide_sheet)
    #         logging.info('輸出 Report_Target/QC 輸出 RL 值')
    #     if vendor.upper()=='AMD':
    #         #### ---- 輸出 OPT_Status
    #         export_report.__table(df_opt_status,report_filename,'OPT_Status',b_hide_sheet)
    #         logging.info('輸出 OPT_Status')
    #     export_report.__table(df_opt_summuary,report_filename,'OPT_SummaryCap',b_hide_sheet)
    #     logging.info('輸出 OPT_SummaryCap')
    #     #### ---- 輸出 OPT_TotalCap
    #     export_report.__table(df_opt_totalcap,report_filename,'OPT_TotalCap',b_hide_sheet)
    #     logging.info('輸出 OPT_TotalCap')
    #     #### ---- 輸出 OPT_RowData
    #     export_report.__table(opt_raw_data,report_filename,'OPT_RowData',b_hide_sheet)
    #     logging.info('輸出 OPT_RowData')
    # if result == 'Pass':
    #     #### ---- Result_PDN 
    #     export_report.__result_image_to_worksheet(opt_img_list,report_filename,'OPI_15%',False,vendor)
    #     logging.info('輸出 OPT result')
    # logging.info(f'OPT Status:{result}')
    # # ! ---- cost report -----------------------------------------------------------
    # logging.info('輸出 cost report')
    # #### ---- NetName
    # export_report.__table(df_netname,cost_report_filename,'NetName',b_hide_sheet)
    # #### ---- ORI_TotalCap
    # export_report.__table(df_ori_table,cost_report_filename,'ORI_TotalCap',b_hide_sheet)
    # if result:
    #     #### ---- OPT_Status
    #     export_report.__table(df_opt_status,cost_report_filename,'OPT_Status',b_hide_sheet)
    #     #### ---- OPT_TotalCap
    #     export_report.__table(df_opt_totalcap,cost_report_filename,'OPT_TotalCap',b_hide_sheet)
    #     #### ---- NOPT_TotalCap
    #     export_report.__table(df_nopt_totalcap,cost_report_filename,'NOPT_TotalCap',b_hide_sheet)
    #     logging.info('輸出 cost report 完成')

    # # ! 讀取cost資料準備進DB
    # # REMARK
    # export_report.__read_cost_summary(cost_report_filename)

    # # copy Excel 到資料夾位置
    # shutil.copy2(report_filename, output_path)
    # os.remove(report_filename) # 刪除檔案
    # shutil.copy2(cost_report_filename, output_path)
    # os.remove(cost_report_filename) # 刪除檔案
